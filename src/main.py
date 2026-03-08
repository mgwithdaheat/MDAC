# src/main.py
# Compatibil Python 3.8+ + export CSV si XLSX (CNP pastrat ca text in Excel)

import re
from datetime import date
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd
import pdfplumber
from tqdm import tqdm

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from pdf2image import convert_from_path
import pytesseract

# pentru formatare Excel
from openpyxl import load_workbook


# ---------------------------
# Paths / Config
# ---------------------------
BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = BASE_DIR / "data" / "input_pdfs"
OUTPUT_DIR = BASE_DIR / "data" / "output"

CNP_RE = re.compile(r"\b[1-8]\d{12}\b")


# ---------------------------
# PDF text extraction
# ---------------------------
def extract_text_pdfplumber(pdf_path: Path) -> str:
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts).strip()


def extract_text_ocr(pdf_path: Path) -> str:
    images = convert_from_path(str(pdf_path), dpi=250)
    parts = []
    for img in images:
        parts.append(pytesseract.image_to_string(img, lang="ron+eng"))
    return "\n".join(parts).strip()


def get_pdf_text(pdf_path: Path) -> Tuple[str, str]:
    text = extract_text_pdfplumber(pdf_path)
    if len(text) < 80:
        return extract_text_ocr(pdf_path), "ocr"
    return text, "text"


# ---------------------------
# Field extraction
# ---------------------------
def extract_cnp(text: str) -> Optional[str]:
    m = CNP_RE.search(text)
    return m.group(0) if m else None


# ---------------------------
# Enrichment from CNP
# ---------------------------
def cnp_to_sex(cnp: str) -> Optional[str]:
    if not cnp or len(cnp) != 13:
        return None
    # simplificat
    return "M" if cnp[0] in "1357" else "F"


def cnp_to_birthdate(cnp: str) -> Optional[date]:
    if not cnp or len(cnp) != 13:
        return None

    s = int(cnp[0])
    yy, mm, dd = int(cnp[1:3]), int(cnp[3:5]), int(cnp[5:7])

    if s in (1, 2):
        century = 1900
    elif s in (5, 6, 7, 8, 3, 4):
        # pentru proiect e suficient (nu diferentiem fin 1800/2000 la toate cazurile)
        century = 2000 if s in (5, 6, 7, 8) else 1800
    else:
        return None

    try:
        return date(century + yy, mm, dd)
    except ValueError:
        return None


def age_from_birthdate(bd: Optional[date]) -> Optional[int]:
    if not bd:
        return None
    today = date.today()
    return today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))


# ---------------------------
# Report PDF
# ---------------------------
def write_report_pdf(path: Path, stats: dict, author: str):
    path.parent.mkdir(parents=True, exist_ok=True)

    c = canvas.Canvas(str(path), pagesize=A4)
    y = A4[1] - 60
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Raport analiza declaratii (OCR + extractie)")
    y -= 30
    c.setFont("Helvetica", 11)
    c.drawString(50, y, f"Autor: {author}")
    y -= 25

    for k, v in stats.items():
        c.drawString(50, y, f"{k}: {v}")
        y -= 16
        if y < 80:
            c.showPage()
            y = A4[1] - 60
            c.setFont("Helvetica", 11)

    c.showPage()
    c.save()


# ---------------------------
# Excel helpers (CNP ca text)
# ---------------------------
def export_xlsx_with_text_column(df: pd.DataFrame, out_xlsx: Path, text_columns=None) -> None:
    """
    Exporta DataFrame in XLSX si forteaza anumite coloane sa fie TEXT in Excel
    (evita notatia stiintifica de forma 1.64112E+12 pentru CNP).
    """
    if text_columns is None:
        text_columns = ["cnp"]

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # Scriem initial cu pandas
    df.to_excel(out_xlsx, index=False)

    # Apoi formatam coloanele ca text cu openpyxl
    wb = load_workbook(out_xlsx)
    ws = wb.active

    # Header row = 1
    headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

    for col_name in text_columns:
        if col_name not in headers:
            continue
        col_idx = headers[col_name]
        # setam format text '@' + convertim explicit la string
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is None:
                continue
            cell.number_format = "@"
            cell.value = str(cell.value)

    wb.save(out_xlsx)


# ---------------------------
# Main
# ---------------------------
def main(author_name: str = "Numele Tau") -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(INPUT_DIR.glob("*.pdf"))
    if not pdf_files:
        print(f"⚠️ Nu exista PDF-uri in: {INPUT_DIR}")
        print("Ruleaza mai intai: python src/generate_pdfs.py (sau pune PDF-uri in data/input_pdfs)")
        return

    rows = []
    for pdf in tqdm(pdf_files, desc="Procesare PDF"):
        try:
            text, method = get_pdf_text(pdf)
        except Exception as e:
            rows.append({
                "file": pdf.name,
                "method": "error",
                "cnp": None,
                "sex": None,
                "birthdate": None,
                "age": None,
                "error": str(e),
            })
            continue

        cnp = extract_cnp(text)
        bd = cnp_to_birthdate(cnp) if cnp else None

        rows.append({
            "file": pdf.name,
            "method": method,
            "cnp": cnp,  # il pastram ca string
            "sex": cnp_to_sex(cnp) if cnp else None,
            "birthdate": bd.isoformat() if bd else None,
            "age": age_from_birthdate(bd),
            "error": None,
        })

    df = pd.DataFrame(rows)

    # --- Export CSV (pentru compatibilitate) ---
    # OPTIONAL: prefix apostrof doar la export CSV ca Excel sa-l vada text
    df_csv = df.copy()
    df_csv["cnp"] = df_csv["cnp"].apply(lambda x: f"'{x}" if pd.notna(x) else x)

    out_csv = OUTPUT_DIR / "extracted.csv"
    df_csv.to_csv(out_csv, index=False, encoding="utf-8-sig")

    # --- Export XLSX (recomandat pentru Excel) ---
    out_xlsx = OUTPUT_DIR / "extracted.xlsx"
    export_xlsx_with_text_column(df, out_xlsx, text_columns=["cnp"])

    # Statistici
    female_count = int((df["sex"] == "F").sum())
    over_50_count = int((df["age"].fillna(-1) > 50).sum())

    stats = {
        "Total documente": int(len(df)),
        "Fisiere OCR": int((df["method"] == "ocr").sum()),
        "Femei (din CNP)": female_count,
        "Peste 50 ani": over_50_count,
        "Erori la procesare": int((df["method"] == "error").sum()),
    }

    out_report = OUTPUT_DIR / "report.pdf"
    write_report_pdf(out_report, stats, author_name)

    print("\n✅ Gata.")
    print(f"CSV (Excel-safe cu apostrof): {out_csv}")
    print(f"XLSX (CNP text): {out_xlsx}")
    print(f"Raport PDF: {out_report}")


if __name__ == "__main__":
    main(author_name="Numele Tau")
